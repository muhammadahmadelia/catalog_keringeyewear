import os
import sys
from time import sleep
import json
import glob
import threading
from models.store import Store
from models.brand import Brand
from models.product import Product
from models.metafields import Metafields
from models.variant import Variant

from datetime import datetime
from selenium import webdriver
import chromedriver_autoinstaller
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from bs4 import BeautifulSoup
import requests
import unidecode

from openpyxl import Workbook
from openpyxl.drawing.image import Image as Imag
from openpyxl.utils import get_column_letter
from PIL import Image

class myScrapingThread(threading.Thread):
    def __init__(self, threadID: int, name: str, obj, brand: str, glasses_type: str, product_number: str, product_url: str, headers: dict) -> None:
        threading.Thread.__init__(self)
        self.threadID = threadID
        self.name = name
        self.brand = brand
        self.glasses_type = glasses_type
        self.product_number = product_number
        self.product_url = product_url
        self.headers = headers
        self.obj = obj
        self.status = 'in progress'
        pass

    def run(self):
        self.obj.scrape_product(self.brand, self.glasses_type, self.product_number, self.product_url, self.headers)
        self.status = 'completed'

    def active_threads(self):
        return threading.activeCount()

class Keringeyewear_Scraper:
    def __init__(self, DEBUG: bool, result_filename: str, logs_filename: str) -> None:
        self.DEBUG = DEBUG
        self.data = []
        self.result_filename = result_filename
        self.logs_filename = logs_filename
        self.thread_list = []
        self.thread_counter = 0
        self.chrome_options = Options()
        self.chrome_options.add_argument('--disable-infobars')
        self.chrome_options.add_argument("--start-maximized")
        self.chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
        self.args = ["hide_console", ]
        self.browser = webdriver.Chrome(options=self.chrome_options, service_args=self.args)
        pass

    def controller(self, store: Store, brands_with_types: list[dict]) -> None:
        try:
            cookies = ''

            self.browser.get(store.link)
            self.wait_until_browsing()
            self.accept_cookies()

            if self.login(store.username, store.password):

                for brand_with_type in brands_with_types:
                    brand: str = brand_with_type['brand']
                    brand_code: str = brand_with_type['code']
                    print(f'Brand: {brand}')

                    for glasses_type in brand_with_type['glasses_type']:

                        ActionChains(self.browser).move_to_element(self.browser.find_element(By.CSS_SELECTOR, 'li[class="col-md-auto plp-menu"]')).perform()
                        sleep(0.8)

                        brand_url = self.get_brand_url(brand, glasses_type)
                        self.open_new_tab(brand_url)

                        self.wait_until_loading()
                        self.load_all_products()
                        
                        total_products = self.get_total_products()
                        scraped_products = 0

                        print(f'Type: {glasses_type} | Total products: {total_products}')
                        start_time = datetime.now()
                        print(f'Start Time: {start_time.strftime("%A, %d %b %Y %I:%M:%S %p")}')

                        self.printProgressBar(scraped_products, total_products, prefix = 'Progress:', suffix = 'Complete', length = 50)

                        if not cookies: cookies = self.get_cookies()
                        headers = self.get_headers(cookies, brand_url)

                        for product_div in self.browser.find_elements(By.XPATH, '//div[@class="product-item space purchasable-plp set-border "]'):
                            try:
                                scraped_products += 1
                                # ActionChains(self.browser).move_to_element(product_div).perform()

                                product_url = product_div.find_element(By.CSS_SELECTOR, 'a[class^="name"]').get_attribute('data-producturl')
                                if 'https://my.keringeyewear.com' not in product_url: product_url = f'https://my.keringeyewear.com{product_url}'
                                product_number = str(product_div.find_element(By.CSS_SELECTOR, 'a[class^="name"] > span').text).strip()

                                self.create_thread(brand, glasses_type, product_number, product_url, headers)
                                sleep(0.5)
                                if self.thread_counter >= 40: 
                                    self.wait_for_thread_list_to_complete()
                                    self.save_to_json(self.data)

                                self.printProgressBar(scraped_products, total_products, prefix = 'Progress:', suffix = 'Complete', length = 50)

                            except Exception as e:
                                self.print_logs(f'Exception in product_div loop: {e}')
                                if self.DEBUG: print(f'Exception in product_div loop: {e}')

                        self.wait_for_thread_list_to_complete()
                        self.save_to_json(self.data)
                        self.close_last_tab()

                        end_time = datetime.now()
                        print(f'End Time: {end_time.strftime("%A, %d %b %Y %I:%M:%S %p")}')
                        print('Duration: {}\n'.format(end_time - start_time))

                        ActionChains(self.browser).move_to_element(self.browser.find_element(By.CSS_SELECTOR, 'div[class="logo"]')).perform()
                        sleep(0.5)
            else: print(f'Failed to login \nURL: {store.link}\nUsername: {str(store.username)}\nPassword: {str(store.password)}')

        except Exception as e:
            if self.DEBUG: print(f'Exception in Keringeyewear_Scraper controller: {e}')
            self.print_logs(f'Exception in Keringeyewear_Scraper controller: {e}')
        finally: 
            self.browser.quit()
            self.wait_for_thread_list_to_complete()
            self.save_to_json(self.data)

    def accept_cookies(self) -> None:
        try:
            self.wait_until_element_found(40, 'xpath', '//button[@id="onetrust-accept-btn-handler"]')
            sleep(3)
            self.browser.find_element(By.XPATH, '//button[@id="onetrust-accept-btn-handler"]').click()
            sleep(2)
        except Exception as e:
            if self.DEBUG: print(f'Exception in accept_cookies: {str(e)}')
            self.print_logs(f'Exception in accept_cookies: {str(e)}')

    def get_cookie_value(self, cookie_name):
        cookie_value = ''
        try:
            for browser_cookie in self.browser.get_cookies():
                if browser_cookie['name'] == cookie_name:
                    cookie_value = browser_cookie['value']
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_cookie_value: {e}')
            self.print_logs(f'Exception in get_cookie_value: {e}')
        finally: return cookie_value

    def wait_until_element_found(self, wait_value: int, type: str, value: str) -> bool:
        flag = False
        try:
            if type == 'id':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.ID, value)))
                flag = True
            elif type == 'xpath':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.XPATH, value)))
                flag = True
            elif type == 'css_selector':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.CSS_SELECTOR, value)))
                flag = True
            elif type == 'class_name':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.CLASS_NAME, value)))
                flag = True
            elif type == 'tag_name':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.TAG_NAME, value)))
                flag = True
        except: pass
        finally: return flag

    def login(self, username: str, password: str) -> bool:
        login_flag = False
        try:
            if self.wait_until_element_found(20, 'xpath', '//input[@name="j_username"]'):
                self.browser.find_element(By.XPATH, '//input[@name="j_username"]').send_keys(username)
                if self.wait_until_element_found(20, 'xpath', '//input[@name="j_password"]'):
                    self.browser.find_element(By.XPATH, '//input[@name="j_password"]').send_keys(password)
                    if self.wait_until_element_found(20, 'xpath', '//button[@name="login"]'):
                        self.browser.find_element(By.XPATH, '//button[@name="login"]').click()
                        for _ in range(0, 50):
                            if self.browser.current_url == 'https://my.keringeyewear.com/en/':
                                login_flag = True
                                break
                            else: sleep(0.3)
        except Exception as e:
            if self.DEBUG: print(f'Exception in login: {str(e)}')
            self.print_logs(f'Exception in login: {str(e)}')
        finally: return login_flag

    def wait_until_browsing(self) -> None:
        while True:
            try:
                state = self.browser.execute_script('return document.readyState; ')
                if 'complete' == state: break
                else: sleep(0.2)
            except: pass

    def wait_until_loading(self) -> None:
        while True:
            try:
                style_class = self.browser.find_element(By.XPATH, '//div[@id="spinner"]').get_attribute('style')
                if 'display: none;' in style_class: break
                else: sleep(0.3)
            except: pass

    def is_xpath_found(self, xpath: str) -> bool:
        try:
            self.browser.find_element(By.XPATH, xpath)
            return True
        except: return False

    def wait_until_xpath_found(self, xpath: str) -> bool:
        for _ in range(0, 100):
            try:
                self.browser.find_element(By.XPATH, xpath)
                return True
            except: sleep(0.5)
        return False

    def wait_until_css_selector_found(self, css_selector: str) -> bool:
        for _ in range(0, 100):
            try:
                self.browser.find_element(By.CSS_SELECTOR, css_selector)
                return True
            except: sleep(0.5)
        return False

    def open_new_tab(self, url: str) -> None:
        # open category in new tab
        self.browser.execute_script('window.open("'+str(url)+'","_blank");')
        self.browser.switch_to.window(self.browser.window_handles[len(self.browser.window_handles) - 1])
        self.wait_until_browsing()
    
    def close_last_tab(self) -> None:
        self.browser.close()
        self.browser.switch_to.window(self.browser.window_handles[len(self.browser.window_handles) - 1])

    def load_all_products(self) -> None:
        total_products, found_products = 0, 0
        while True:
            try:
                if total_products == 0:
                    try: total_products = int(str(self.browser.find_element(By.CSS_SELECTOR, 'div[class*="col-md-11 kering-gray"]').text).strip().split(' ')[0])
                    except: pass
                    # print(f'Total Products: {total_products}')
                try: found_products = len(self.browser.find_elements(By.XPATH, '//div[@class="product-item space purchasable-plp set-border "]'))
                except: pass
                # print(total_products, found_products)
                if int(total_products) == int(found_products): break
                else:
                    if self.is_xpath_found('//div[@class="col-md-12 show-more-button"]'):
                        # button = self.browser.find_element(By.XPATH, '//div[@class="col-md-12 show-more-button"]')
                        # actions = ActionChains(self.browser)
                        # actions.move_to_element(button).perform()
                        # button.click()
                        self.browser.find_element(By.XPATH, '//div[@class="col-md-12 show-more-button"]').click()
                        # self.wait_until_browsing()
                        self.wait_until_loading()
                        sleep(0.8)
                    else: break
            except: pass

    def get_product_data(self, product_number: str, soup: BeautifulSoup) -> list[dict]:
        products_data = []
        try:
            for div in soup.select('div[class*="variants"] > div[class^="product-item space purchasable-plp set-border"]'):
                product_url, number, frame_code, product_size, frame_color, lens_color = '', '', '', '', '', ''
                product_url = div.select_one('form[class="js-product-page"]').get('action')
                if 'https://my.keringeyewear.com' not in product_url: product_url = f'https://my.keringeyewear.com{product_url}'
                text = str(div.select_one('div[class="col-md-12 product-description"] > div[class="details brand"] > a').text).strip()
                number = product_number.strip().upper()
                frame_code = str(text).split('-')[-1].strip().upper()
                frame_code = text.lower().replace(str(number).strip().lower(), '').strip().upper()
                if frame_code[0] == '-': frame_code = frame_code[1:]
                for details_div in div.select('div[class="col-md-12 product-description"] > div[class="details counter-variant"]'):
                    if 'CALIBERS:' in str(details_div.text).strip().upper():
                        product_size = str(details_div.find('span').text).strip()
                    elif 'FRONT:' in str(details_div.text).strip().upper():
                        frame_color = str(details_div.find('span').text).strip()
                    elif 'LENS:' in str(details_div.text).strip().upper():
                        lens_color = str(details_div.find('span').text).strip()

                json_data = {
                    'product_url': product_url,
                    'number': number,
                    'frame_code': frame_code,
                    'frame_color': frame_color,
                    'lens_color': lens_color,
                    'product_size': product_size
                }
                if json_data not in products_data: products_data.append(json_data)
           
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_product_data: {str(e)}')
            self.print_logs(f'Exception in get_product_data: {str(e)}')
        finally: return products_data

    def save_to_json(self, products: list[Product]):
        try:
            json_products = []
            for product in products:
                json_varinats = []
                for index, variant in enumerate(product.variants):
                    json_varinat = {
                        'position': (index + 1), 
                        'title': variant.title, 
                        'sku': variant.sku, 
                        'inventory_quantity': variant.inventory_quantity,
                        'found_status': variant.found_status,
                        'listing_price': variant.listing_price, 
                        'wholesale_price': variant.wholesale_price,
                        'barcode_or_gtin': variant.barcode_or_gtin,
                        'size': variant.size,
                        'weight': variant.weight
                    }
                    json_varinats.append(json_varinat)
                json_product = {
                    'brand': product.brand, 
                    'number': product.number, 
                    'name': product.name, 
                    'frame_code': product.frame_code, 
                    'frame_color': product.frame_color, 
                    'lens_code': product.lens_code, 
                    'lens_color': product.lens_color, 
                    'status': product.status, 
                    'type': product.type, 
                    'url': product.url, 
                    'metafields': [
                        { 'key': 'for_who', 'value': product.metafields.for_who },
                        { 'key': 'product_size', 'value': product.metafields.product_size }, 
                        { 'key': 'lens_material', 'value': product.metafields.lens_material }, 
                        { 'key': 'lens_technology', 'value': product.metafields.lens_technology }, 
                        { 'key': 'frame_material', 'value': product.metafields.frame_material }, 
                        { 'key': 'frame_shape', 'value': product.metafields.frame_shape },
                        { 'key': 'gtin1', 'value': product.metafields.gtin1 }, 
                        { 'key': 'img_url', 'value': product.metafields.img_url },
                        { 'key': 'img_360_urls', 'value': product.metafields.img_360_urls }
                    ],
                    'variants': json_varinats
                }
                json_products.append(json_product)
            
           
            with open(self.result_filename, 'w') as f: json.dump(json_products, f)
            
        except Exception as e:
            if self.DEBUG: print(f'Exception in save_to_json: {e}')
            self.print_logs(f'Exception in save_to_json: {e}')
    
    def get_brand_url(self, brand: str, glasses_type: str) -> str:
        brand_url = ''
        try:
            for a_tag in self.browser.find_elements(By.CSS_SELECTOR, 'div[class*="menu-open brands"] > div[class^="col-md-2"] > a'):
                            
                if str(brand).strip().lower() == unidecode.unidecode(str(a_tag.text).strip()).lower():

                    brand_url = a_tag.get_attribute("href")
                    if glasses_type == 'Sunglasses': brand_url = str(brand_url).strip().replace('&type=Style', '%3AarticleType%3ASun&target=product&type=Style#')
                    elif glasses_type == 'Eyeglasses': brand_url = str(brand_url).strip().replace('&type=Style', '%3AarticleType%3AOptical&target=product&type=Style#')
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_brand_url: {e}')
            self.print_logs(f'Exception in get_brand_url: {e}')
        finally: return brand_url

    def get_total_products(self) -> int:
        total_products = 0
        try:
            total_products = len(self.browser.find_elements(By.XPATH, '//div[@class="product-item space purchasable-plp set-border "]'))
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_brand_url: {e}')
            self.print_logs(f'Exception in get_brand_url: {e}')
        finally: return total_products

    def get_cookies(self) -> str:
        cookies = ''
        try:
            cookies = f"HYBRIS-SRV={self.get_cookie_value('HYBRIS-SRV')}; JSESSIONID={self.get_cookie_value('JSESSIONID')}; anonymous-consents={self.get_cookie_value('anonymous-consents')}; "
            cookies += f"cookie-notification={self.get_cookie_value('cookie-notification')}; ROUTE={self.get_cookie_value('ROUTE')}; ASLBSA={self.get_cookie_value('ASLBSA')}; ASLBSACORS={self.get_cookie_value('ASLBSACORS')}; "
            cookies += f"__utma={self.get_cookie_value('__utma')}; __utmc={self.get_cookie_value('__utmc')}; __utmz={self.get_cookie_value('__utmz')}; __utmt={self.get_cookie_value('__utmt')}; "
            cookies += f"OptanonAlertBoxClosed={self.get_cookie_value('OptanonAlertBoxClosed')}; _ga={self.get_cookie_value('_ga')}; _gid={self.get_cookie_value('_gid')}; securityToken={self.get_cookie_value('securityToken')}; "
            cookies += f"acceleratorSecureGUID={self.get_cookie_value('acceleratorSecureGUID')}; UPSELLsun3={self.get_cookie_value('UPSELLsun3')}: UPSELLoptical3={self.get_cookie_value('UPSELLoptical3')}; _gat_gtag_UA_72952013_2=1; "
            cookies += f"__utmb={self.get_cookie_value('__utmb')}; OptanonConsent={self.get_cookie_value('OptanonConsent')}"
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_cookies: {e}')
            self.print_logs(f'Exception in get_cookies: {e}')
        finally: return cookies

    def get_headers(self, cookies: str, brand_url: str) -> dict:
        return {
                'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
                'accept-encoding': 'gzip, deflate, br',
                'accept-language': 'en-US,en;q=0.9',
                'cache-control': 'max-age=0',
                'cookie': cookies,
                'referer': brand_url,
                'sec-ch-ua': '"Chromium";v="104", " Not A;Brand";v="99", "Google Chrome";v="104"',
                'sec-ch-ua-mobile': '?0',
                'sec-ch-ua-platform': '"Windows"',
                'sec-fetch-dest': 'document',
                'sec-fetch-mode': 'navigate',
                'sec-fetch-site': 'same-origin',
                'sec-fetch-user': '?1',
                'upgrade-insecure-requests': '1',
                'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.5112.101 Safari/537.36'
            }

    def scrape_product(self, brand: str, glasses_type: str, product_number: str, product_url: str, headers: dict):
        try:
            URL = product_url
            response = self.get_response(URL, headers)
            if response and response.status_code == 200:
                
                soup = BeautifulSoup(response.text, 'lxml')
                products_data = self.get_product_data(product_number, soup)
                
                for product_data in products_data:
                    try:
                        product = Product()
                        product.brand = brand
                        product.number = product_data['number']
                        product.frame_code = product_data['frame_code']
                        product.frame_color = product_data['frame_color']
                        product.lens_color = product_data['lens_color']
                        product.status = 'active'
                        product.type = glasses_type
                        product.url = product_data['product_url']

                        if product.url not in URL:
                            URL = product.url
                            response = self.get_response(URL, headers)
                            if response and response.status_code == 200:
                                soup = BeautifulSoup(response.text, 'lxml')
                            

                        metafields = self.scrape_product_metafields(product_data, product, soup)

                        variant = self.scrape_product_variant(product, metafields, soup)
                        
                        metafields.gtin1 = variant.barcode_or_gtin
                        product.metafields = metafields
                        product.variants = variant

                        self.data.append(product)
                    except Exception as e:
                        if self.DEBUG: print(f'Exception in product: {e}')
                        self.print_logs(f'Exception in product: {e}')
        except Exception as e:
            if self.DEBUG: print(f'Exception in scrape_product: {e}')
            self.print_logs(f'Exception in scrape_product: {e}')

    def get_response(self, url: str, headers: dict):
        response = ''
        try:
            for _ in range(0, 10):
                try:
                    response = requests.get(url=url, headers=headers)
                    if response.status_code == 200: break
                except: sleep(0.1)
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_response: {e}')
            self.print_logs(f'Exception in get_response: {e}')
        finally: return response

    def scrape_product_metafields(self, product_data: dict, product: Product, soup: BeautifulSoup) -> Metafields:
        metafields = Metafields()
        try:
            if str(product_data['product_size']).strip() != '-0-0':
                metafields.product_size = str(product_data['product_size']).strip()

            try:
                for div in soup.select('div[id="kering-product-characteristics"] > div[id="kering-product-characteristics-collapsable"] > div[class="col-sm-12 col-xs-12"]'):
                    if str(div.select_one('span[class*="characteristics-title"]').text).strip().lower() == 'gender':
                        metafields.for_who = str(div.find_all('span')[1].text).strip().title()
                        break
            except Exception as e: 
                if self.DEBUG: print(f'Exception in metafields.for_who: {e}')
                else: sleep(0.15)

            try:
                for div in soup.select('div[id="kering-product-characteristics"] > div[id="kering-product-characteristics-collapsable"] > div[class="col-sm-6 col-xs-12"]'):
                    if str(div.select_one('span[class*="characteristics-title"]').text).strip().lower() == 'temple main':
                        metafields.frame_material = str(div.find_all('span')[1].text).strip().title()
                    if str(div.select_one('span[class*="characteristics-title"]').text).strip().lower() == 'lens':
                        text = str(div.find_all('span')[1].text).strip().title()
                        if str(product.lens_color).strip().lower() != str(text).strip().lower():
                            metafields.lens_material = text
                    if metafields.frame_material and metafields.lens_material: break
            except Exception as e: 
                if self.DEBUG: print(f'Exception in metafields.frame_material: {e}')
                else: sleep(0.15)

            try:
                img_tag = soup.select_one('div > img[class="lazyOwl"]')
                metafields.img_url = img_tag.get('src') if img_tag else ''
                if 'missing_product_EN_512x512.png' in metafields.img_url: metafields.img_url = ''
            except Exception as e: 
                if self.DEBUG: print(f'Exception in metafields.img_url: {e}')
                else: sleep(0.15)


            # if metafields.img_url:
            #     try:
            #         for img_tag in soup.select('div[class="item"] > img[class="lazyOwl"]'):
            #             metafields.img_360_urls = img_tag.get('src')
            #     except Exception as e: 
            #         if self.DEBUG: print(f'Exception in metafields.img_360_urls: {e}')
            #         else: sleep(0.15)
        except Exception as e:
            if self.DEBUG: print(f'Exception in scrape_product_metafields: {e}')
            self.print_logs(f'Exception in scrape_product_metafields: {e}')
        finally: return metafields

    def scrape_product_variant(self, product: Product, metafields: Metafields, soup: BeautifulSoup) -> Variant:
        variant = Variant()
        try:
            variant.position = len(product.variants) + 1
            if metafields.product_size: 
                variant.title = str(metafields.product_size).split('-')[0].strip()
                variant.size = metafields.product_size
            if variant.title: variant.sku = f'{product.number} {product.frame_code} {variant.title}'
            else: variant.sku = f'{product.number} {product.frame_code}'

            try:
                for span_tag in soup.select('div[class^="srp price-srp"] >span'):
                    if '€' in str(span_tag.text).strip(): 
                        variant.listing_price = str(span_tag.text).strip().replace('€', '').strip()
                        break
            except Exception as e: 
                if self.DEBUG: print(f'Exception in variant.price: {e}')
                else: sleep(0.15)

            try:
                span_tag = soup.select_one('div[class^="whs price-whs"] > span')
                if span_tag and '€' in str(span_tag.text).strip(): 
                    variant.wholesale_price = str(span_tag.text).strip().replace('€', '').strip()
            except Exception as e: 
                if self.DEBUG: print(f'Exception in variant.wholesale_price: {e}')
                else: sleep(0.15)
            variant.found_status = 1

            try:
                if '/available.svg' in soup.select_one('div[class^="package-status"] > img').get('src'):
                    variant.inventory_quantity = 1
                else: variant.inventory_quantity = 0
            except: variant.inventory_quantity = 0
        
            try:
                for div in reversed(soup.select('div[id="kering-product-characteristics"] > div[id="kering-product-characteristics-collapsable"] > div[class="col-sm-6 col-xs-12"]')):
                    # if str(div.select_one('span[class*="characteristics-title"]').text).strip().lower() == 'ean':
                    if str(div.select_one('span[class*="characteristics-title"]').text).strip().lower() == 'upc':
                        variant.barcode_or_gtin = str(div.find_all('span')[1].text).strip()
                        break
            except Exception as e: 
                if self.DEBUG: print(f'Exception in variant.barcode_or_gtin: {e}')
                else: sleep(0.15)
        except Exception as e:
            if self.DEBUG: print(f'Exception in scrape_product_variant: {e}')
            self.print_logs(f'Exception in scrape_product_variant: {e}')
        finally: return variant

    # print logs to the log file
    def print_logs(self, log: str) -> None:
        try:
            with open(self.logs_filename, 'a') as f:
                f.write(f'\n{log}')
        except: pass

    def printProgressBar(self, iteration, total, prefix = '', suffix = '', decimals = 1, length = 100, fill = '█', printEnd = "\r") -> None:
        """
        Call in a loop to create terminal progress bar
        @params:
            iteration   - Required  : current iteration (Int)
            total       - Required  : total iterations (Int)
            prefix      - Optional  : prefix string (Str)
            suffix      - Optional  : suffix string (Str)
            decimals    - Optional  : positive number of decimals in percent complete (Int)
            length      - Optional  : character length of bar (Int)
            fill        - Optional  : bar fill character (Str)
            printEnd    - Optional  : end character (e.g. "\r", "\r\n") (Str)
        """
        percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
        filledLength = int(length * iteration // total)
        bar = fill * filledLength + '-' * (length - filledLength)
        print(f'\r{prefix} |{bar}| {percent}% {suffix}', end = printEnd)
        # Print New Line on Complete
        if iteration == total: 
            print()

    def create_thread(self, brand: str, glasses_type: str, product_number: str, product_url: str, headers: dict):
        thread_name = "Thread-"+str(self.thread_counter)
        self.thread_list.append(myScrapingThread(self.thread_counter, thread_name, self, brand, glasses_type, product_number, product_url, headers))
        self.thread_list[self.thread_counter].start()
        self.thread_counter += 1

    def is_thread_list_complted(self):
        for obj in self.thread_list:
            if obj.status == "in progress":
                return False
        return True

    def wait_for_thread_list_to_complete(self):
        while True:
            result = self.is_thread_list_complted()
            if result: 
                self.thread_counter = 0
                self.thread_list.clear()
                break
            else: sleep(1)

def read_data_from_json_file(DEBUG, result_filename: str):
    data = []
    try:
        files = glob.glob(result_filename)
        if files:
            f = open(files[-1])
            json_data = json.loads(f.read())
            products = []

            for json_d in json_data:
                number, frame_code, brand, img_url, frame_color, lens_color= '', '', '', '', '', ''
                # product = Product()
                brand = json_d['brand']
                number = str(json_d['number']).strip().upper()
                if '/' in number: number = number.replace('/', '-').strip()
                # product.name = str(json_d['name']).strip().upper()
                frame_code = str(json_d['frame_code']).strip().upper()
                if '/' in frame_code: frame_code = frame_code.replace('/', '-').strip()
                frame_color = str(json_d['frame_color']).strip().title()
                # product.lens_code = str(json_d['lens_code']).strip().upper()
                lens_color = str(json_d['lens_color']).strip().title()
                # product.status = str(json_d['status']).strip().lower()
                # product.type = str(json_d['type']).strip().title()
                # product.url = str(json_d['url']).strip()
                # metafields = Metafields()
                
                for json_metafiels in json_d['metafields']:
                    # if json_metafiels['key'] == 'for_who':metafields.for_who = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'product_size':metafields.product_size = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'activity':metafields.activity = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'lens_material':metafields.lens_material = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'graduabile':metafields.graduabile = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'interest':metafields.interest = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'lens_technology':metafields.lens_technology = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'frame_material':metafields.frame_material = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'frame_shape':metafields.frame_shape = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'gtin1':metafields.gtin1 = str(json_metafiels['value']).strip().title()
                    if json_metafiels['key'] == 'img_url':img_url = str(json_metafiels['value']).strip()
                    # elif json_metafiels['key'] == 'img_360_urls':
                    #     value = str(json_metafiels['value']).strip()
                    #     if '[' in value: value = str(value).replace('[', '').strip()
                    #     if ']' in value: value = str(value).replace(']', '').strip()
                    #     if "'" in value: value = str(value).replace("'", '').strip()
                    #     for v in value.split(','):
                    #         metafields.img_360_urls = str(v).strip()
                # product.metafields = metafields
                for json_variant in json_d['variants']:
                    sku, price = '', ''
                    # variant = Variant()
                    # variant.position = json_variant['position']
                    # variant.title = str(json_variant['title']).strip()
                    sku = str(json_variant['sku']).strip().upper()
                    if '/' in sku: sku = sku.replace('/', '-').strip()
                    # variant.inventory_quantity = json_variant['inventory_quantity']
                    # variant.found_status = json_variant['found_status']
                    wholesale_price = str(json_variant['wholesale_price']).strip()
                    listing_price = str(json_variant['listing_price']).strip()
                    # variant.barcode_or_gtin = str(json_variant['barcode_or_gtin']).strip()
                    # variant.size = str(json_variant['size']).strip()
                    # variant.weight = str(json_variant['weight']).strip()
                    # product.variants = variant
                    image_attachment = download_image(img_url)
                    if image_attachment:
                        with open(f'Images/{sku}.jpg', 'wb') as f: f.write(image_attachment)
                        crop_downloaded_image(f'Images/{sku}.jpg')
                    data.append([number, frame_code, frame_color, lens_color, brand, sku, wholesale_price, listing_price])
    except Exception as e:
        if DEBUG: print(f'Exception in read_data_from_json_file: {e}')
        else: pass
    finally: return data

def download_image(url):
    image_attachment = ''
    try:
        headers = {
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'accept-Encoding': 'gzip, deflate, br',
            'accept-Language': 'en-US,en;q=0.9',
            'cache-Control': 'max-age=0',
            'sec-ch-ua': '"Google Chrome";v="95", "Chromium";v="95", ";Not A Brand";v="99"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'document',
            'sec-fetch-mode': 'navigate',
            'sec-fetch-site': 'none',
            'Sec-Fetch-User': '?1',
            'upgrade-insecure-requests': '1',
        }
        counter = 0
        while True:
            try:
                response = requests.get(url=url, headers=headers, timeout=20)
                if response.status_code == 200:
                    # image_attachment = base64.b64encode(response.content)
                    image_attachment = response.content
                    break
                else: print(f'{response.status_code} found for url: {url}')
            except: sleep(0.3)
            counter += 1
            if counter == 10: break
    except Exception as e: print(f'Exception in download_image: {str(e)}')
    finally: return image_attachment

def crop_downloaded_image(filename):
    try:
        im = Image.open(filename)
        width, height = im.size   # Get dimensions
        new_width = 1680
        new_height = 1020
        if width > new_width and height > new_height:
            left = (width - new_width)/2
            top = (height - new_height)/2
            right = (width + new_width)/2
            bottom = (height + new_height)/2
            im = im.crop((left, top, right, bottom))
            # im.save(filename)
            try:
                im.save(filename)
            except:
                rgb_im = im.convert('RGB')
                rgb_im.save(filename)
    except Exception as e: print(f'Exception in crop_downloaded_image: {e}')
    
def saving_picture_in_excel(data: list):
    workbook = Workbook()
    worksheet = workbook.active

    worksheet.cell(row=1, column=1, value='Model Code')
    worksheet.cell(row=1, column=2, value='Lens Code')
    worksheet.cell(row=1, column=3, value='Color Frame')
    worksheet.cell(row=1, column=4, value='Color Lens')
    worksheet.cell(row=1, column=5, value='Brand')
    worksheet.cell(row=1, column=6, value='SKU')
    worksheet.cell(row=1, column=7, value='Wholesale Price')
    worksheet.cell(row=1, column=8, value='Listing Price')
    worksheet.cell(row=1, column=9, value="Image")

    for index, d in enumerate(data):
        new_index = index + 2

        worksheet.cell(row=new_index, column=1, value=d[0])
        worksheet.cell(row=new_index, column=2, value=d[1])
        worksheet.cell(row=new_index, column=3, value=d[2])
        worksheet.cell(row=new_index, column=4, value=d[3])
        worksheet.cell(row=new_index, column=5, value=d[4])
        worksheet.cell(row=new_index, column=6, value=d[5])
        worksheet.cell(row=new_index, column=7, value=d[6])
        worksheet.cell(row=new_index, column=8, value=d[7])

        image = f'Images/{d[-3]}.jpg'
        if os.path.exists(image):
            im = Image.open(image)
            width, height = im.size
            worksheet.row_dimensions[new_index].height = height
            worksheet.add_image(Imag(image), anchor='I'+str(new_index))
            # col_letter = get_column_letter(9)
            # worksheet.column_dimensions[col_letter].width = width
        # print(index, image)

    workbook.save('Keringeyewear Results.xlsx')


DEBUG = True
try:
    pathofpyfolder = os.path.realpath(sys.argv[0])
    # get path of Exe folder
    path = pathofpyfolder.replace(pathofpyfolder.split('\\')[-1], '')
    # download chromedriver.exe with same version and get its path
    if os.path.exists('chromedriver.exe'): os.remove('chromedriver.exe')
    if os.path.exists('Keringeyewear Results.xlsx'): os.remove('Keringeyewear Results.xlsx')

    chromedriver_autoinstaller.install(path)
    if '.exe' in pathofpyfolder.split('\\')[-1]: DEBUG = False
    
    f = open('Keringeyewear start.json')
    json_data = json.loads(f.read())
    f.close()

    brands = json_data['brands']

    
    f = open('requirements/keringeyewear.json')
    data = json.loads(f.read())
    f.close()

    store = Store()
    store.link = data['url']
    store.username = data['username']
    store.password = data['password']
    store.login_flag = True

    result_filename = 'requirements/Keringeyewear Results.json'

    if not os.path.exists('Logs'): os.makedirs('Logs')

    log_files = glob.glob('Logs/*.txt')
    if len(log_files) > 5:
        oldest_file = min(log_files, key=os.path.getctime)
        os.remove(oldest_file)
        log_files = glob.glob('Logs/*.txt')

    scrape_time = datetime.now().strftime('%d-%m-%Y %H-%M-%S')
    logs_filename = f'Logs/Logs {scrape_time}.txt'
    
    Keringeyewear_Scraper(DEBUG, result_filename, logs_filename).controller(store, brands)
    
    for filename in glob.glob('Images/*'): os.remove(filename)
    data = read_data_from_json_file(DEBUG, result_filename)
    os.remove(result_filename)

    saving_picture_in_excel(data)
except Exception as e:
    if DEBUG: print('Exception: '+str(e))
    else: pass